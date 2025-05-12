import os
import sys
from pathlib import Path
import tempfile
import logging
import shutil
import json
import uuid
from venv import logger
from werkzeug.utils import secure_filename
from flask import render_template, request, send_file, flash, jsonify, redirect, url_for, session
from web import app
from web.forms import AssetForm, SearchForm, FileUploadForm
import pandas as pd

# Agregar el directorio raíz al path de manera más robusta
ROOT_DIR = Path(__file__).parent.parent
sys.path.insert(0, str(ROOT_DIR))

# Directorio para almacenar resultados temporales
TEMP_RESULTS_DIR = os.path.join(tempfile.gettempdir(), 'freshservice_results')
os.makedirs(TEMP_RESULTS_DIR, exist_ok=True)

from freshservice import FreshServiceManager

manager = FreshServiceManager()

@app.route('/', methods=['GET'])
def index():
    return render_template('index.html')

@app.route('/search', methods=['GET', 'POST'])
def search():
    # Esta ruta ahora redirige a search_id para mantener compatibilidad
    return redirect(url_for('search_id'))

@app.route('/search_id', methods=['GET', 'POST'])
def search_id():
    asset_form = AssetForm()
    results = None
    logger = logging.getLogger(__name__)
    
    # Si es un nuevo GET, limpiar los resultados anteriores
    if request.method == 'GET':
        if 'search_results_id' in session:
            # Limpiar archivo temporal previo si existe
            try:
                temp_file_path = os.path.join(TEMP_RESULTS_DIR, session['search_results_id'] + '.json')
                if os.path.exists(temp_file_path):
                    os.remove(temp_file_path)
                    logger.info(f"Archivo temporal antiguo eliminado: {temp_file_path}")
            except Exception as e:
                logger.error(f"Error al limpiar archivo temporal: {e}")
                
            session.pop('search_results_id')
    
    if request.method == 'POST' and 'submit_assets' in request.form and asset_form.validate_on_submit():
        # Limpiar resultados anteriores
        if 'search_results_id' in session:
            # Limpiar archivo temporal previo
            try:
                temp_file_path = os.path.join(TEMP_RESULTS_DIR, session['search_results_id'] + '.json')
                if os.path.exists(temp_file_path):
                    os.remove(temp_file_path)
                    logger.info(f"Archivo temporal antiguo eliminado: {temp_file_path}")
            except Exception as e:
                logger.error(f"Error al limpiar archivo temporal: {e}")
                
            session.pop('search_results_id')
            
        # Obtener componentes seleccionados
        components = request.form.getlist('components')
        
        # Procesar búsqueda por IDs
        options = {
            'ids': asset_form.ids.data,
            'exclude': asset_form.exclude.data,
            'components': components if components else None,
            'disable_join': asset_form.disable_join.data,
            'combine_cpu_ram': asset_form.combine_cpu_ram.data,
            'asset_data': True,  # Siempre True para obtener datos básicos
            'include_departments': 'departments' in asset_form.include_info.data,
            'include_user': 'user' in asset_form.include_info.data,
            'include_location': 'location' in asset_form.include_info.data,
            'include_system_os': 'system' in asset_form.include_info.data,
            'include_machine_ip': 'machine_ip' in asset_form.include_info.data,
            'include_machine_mac': 'machine_mac' in asset_form.include_info.data,
            'include_serial_number': 'serial_number' in asset_form.include_info.data,
            'include_description': 'description' in asset_form.include_info.data,
            'verbose': True,
            'all_data': asset_form.all_data.data
        }
        
        try:
            # Procesar los resultados primero
            results = manager.run_and_get_results(options)
            
            if not results:
                flash('No se encontraron resultados', 'warning')
                return render_template('search_id.html', asset_form=asset_form)
            
            # Para mostrar en la web, formatear los resultados
            formatted_results = []
            for result in results:
                formatted_result = {
                    'ID': result.get('display_id'),
                    'Nombre': result.get('name'),
                    'Tipo': result.get('asset_type', 'Desconocido'),
                }
                
                # Información adicional
                if 'department' in result:
                    formatted_result['Departamento'] = result['department']
                    
                # Siempre incluir campos de usuario (vacíos si no hay usuario)
                if 'user' in result:
                    # Extraer todos los campos del usuario y formatearlos individualmente
                    user_data = result['user']
                    formatted_result.update({
                        'Usuario': f"{user_data.get('first_name', '')} {user_data.get('last_name', '')}",
                        'Email': user_data.get('primary_email', ''),
                        'Departamento Usuario': user_data.get('department_names', ''),
                        'Cargo': user_data.get('job_title', ''),
                        'Teléfono Trabajo': user_data.get('work_phone_number', ''),
                        'Teléfono Móvil': user_data.get('mobile_phone_number', '')
                    })
                else:
                    # Añadir campos de usuario vacíos para mantener la estructura
                    formatted_result.update({
                        'Usuario': '',
                        'Email': '',
                        'Departamento Usuario': '',
                        'Cargo': '',
                        'Teléfono Trabajo': '',
                        'Teléfono Móvil': ''
                    })
                    
                if 'location' in result:
                    formatted_result['Ubicación'] = result['location']
                    
                if 'system_os' in result:
                    formatted_result['Sistema Operativo'] = result['system_os']
                    
                if 'machine_ip' in result:
                    formatted_result['IP'] = result['machine_ip']
                    
                if 'machine_mac' in result:
                    formatted_result['MAC'] = result['machine_mac']
                    
                if 'serial_number' in result:
                    formatted_result['Número de Serie'] = result['serial_number']
                    
                if 'description' in result:
                    formatted_result['Descripción'] = result['description']

                # Componentes (CPU + RAM)
                if result.get('component_type') == 'CPU + RAM':
                    formatted_result.update({
                        'CPU Modelo': result.get('cpu_model'),
                        'CPU Núcleos': result.get('cpu_cores'),
                        'CPU Velocidad': result.get('cpu_speed'),
                        'RAM Capacidad': result.get('ram_capacity'),
                        'RAM Velocidad': result.get('ram_speed'),
                        'RAM Tipo': result.get('ram_memory_type')
                    })
                # Componentes separados
                else:
                    if 'memory_capacity' in result:
                        formatted_result.update({
                            'Memoria RAM': f"{result.get('memory_capacity')}",
                            'Velocidad RAM': result.get('memory_speed'),
                            'Tipo RAM': result.get('memory_type')
                        })
                    if 'cpu_model' in result:
                        formatted_result.update({
                            'CPU Modelo': result.get('cpu_model'),
                            'CPU Núcleos': result.get('cpu_cores'),
                            'CPU Velocidad': result.get('cpu_speed')
                        })
                        
                formatted_results.append(formatted_result)
            
            # Generar un ID único para esta sesión de resultados
            results_id = str(uuid.uuid4())
            
            # Guardar los resultados formateados en un archivo temporal
            temp_file_path = os.path.join(TEMP_RESULTS_DIR, results_id + '.json')
            with open(temp_file_path, 'w', encoding='utf-8') as f:
                json.dump(formatted_results, f)
                
            # Guardar solo el ID en la sesión
            session['search_results_id'] = results_id
            logger.info(f"Resultados guardados en archivo temporal: {temp_file_path}")
            print(f"Resultados guardados en archivo temporal: {temp_file_path}")
            
            return render_template('search_id.html', 
                                asset_form=asset_form,
                                results=formatted_results)

        except Exception as e:
            logger.error(f"Error al procesar la solicitud: {e}")
            flash(f'Error al procesar la solicitud: {str(e)}', 'error')
            return render_template('search_id.html', asset_form=asset_form)
    
    return render_template('search_id.html', asset_form=asset_form)

@app.route('/download_excel', methods=['POST'])
def download_excel():
    """Ruta dedicada para la descarga de Excel desde la búsqueda por ID"""
    logger = logging.getLogger(__name__)
    
    try:
        # Recuperar ID de resultados de la sesión
        results_id = session.get('search_results_id')
        
        # Registra información detallada
        logger.info(f"Exportando resultados de search_id. ID de resultados: {results_id}")
        print(f"Exportando resultados de search_id. ID de resultados: {results_id}")
        
        if not results_id:
            flash('No hay resultados para descargar', 'warning')
            logger.warning("No hay ID de resultados en la sesión")
            return redirect(url_for('search_id'))
            
        # Cargar resultados desde el archivo temporal
        temp_file_path = os.path.join(TEMP_RESULTS_DIR, results_id + '.json')
        if not os.path.exists(temp_file_path):
            flash('No se encontraron los datos de resultados', 'warning')
            logger.warning(f"Archivo temporal no encontrado: {temp_file_path}")
            return redirect(url_for('search_id'))
            
        # Cargar resultados del archivo
        with open(temp_file_path, 'r', encoding='utf-8') as f:
            try:
                results = json.load(f)
                logger.info(f"Archivo de resultados cargado: {len(results)} registros")
                print(f"Archivo de resultados cargado: {len(results)} registros")
            except json.JSONDecodeError:
                flash('Error al leer los datos de resultados', 'error')
                logger.error(f"Error al decodificar JSON del archivo: {temp_file_path}")
                return redirect(url_for('search_id'))
        
        # Verificar que hay datos para exportar
        if not results or len(results) == 0:
            flash('Los resultados están vacíos', 'warning')
            logger.warning("La lista de resultados está vacía")
            return redirect(url_for('search_id'))
            
        # Log para depuración
        logger.info(f"Encontrados {len(results)} resultados para exportar")
        print(f"Encontrados {len(results)} resultados para exportar")
        
        # Crear archivo temporal en un directorio temporal
        temp_dir = tempfile.mkdtemp()
        filename = secure_filename(request.form.get('filename') or 'output.xlsx')
        if not filename.endswith('.xlsx'):
            filename += '.xlsx'
            
        output_file = os.path.join(temp_dir, 'temp.xlsx')
        logger.info(f"Archivo temporal creado: {output_file}")
        print(f"Archivo temporal creado: {output_file}")
        
        try:
            # Procesar los resultados por lotes para manejar grandes volúmenes
            # Convertir a DataFrame y asegurar orden correcto de columnas
            df = pd.DataFrame(results)
            
            # Asegurar que ID esté primero, seguido de otras columnas importantes
            priority_columns = ['ID', 'Nombre', 'Tipo', 'Departamento', 'Usuario', 'Email']
            all_columns = df.columns.tolist()
            
            # Filtrar las columnas prioritarias que existen en el dataframe
            existing_priority_columns = [col for col in priority_columns if col in all_columns]
            
            # Obtener las columnas restantes que no son prioritarias
            remaining_columns = [col for col in all_columns if col not in existing_priority_columns]
            
            # Reordenar el dataframe con la nueva orden de columnas
            df = df[existing_priority_columns + remaining_columns]
            
            logger.info(f"DataFrame preparado: {len(df)} filas, {len(df.columns)} columnas")
            print(f"DataFrame preparado: {len(df)} filas, {len(df.columns)} columnas")
            
            # Exportar DataFrame a Excel de forma optimizada
            try:
                # Método 1: Exportación directa
                df.to_excel(output_file, index=False, engine='openpyxl')
                excel_ok = True
                logger.info("Exportación directa completada")
            except Exception as e:
                logger.warning(f"Exportación directa fallida: {str(e)}")
                try:
                    # Método 2: Exportación por lotes
                    writer = pd.ExcelWriter(output_file, engine='openpyxl')
                    
                    # Exportar en lotes de 1000 filas para reducir uso de memoria
                    batch_size = 1000
                    for i in range(0, len(df), batch_size):
                        if i == 0:
                            df.iloc[i:i+batch_size].to_excel(writer, index=False, sheet_name='Sheet1')
                        else:
                            df.iloc[i:i+batch_size].to_excel(writer, index=False, sheet_name=f'Sheet{(i//batch_size)+1}')
                    
                    writer.close()
                    excel_ok = True
                    logger.info("Exportación por lotes completada")
                except Exception as batch_error:
                    logger.error(f"Error en exportación por lotes: {str(batch_error)}")
                    excel_ok = False
            
            # Aplicar formato adicional si se pudo exportar
            if excel_ok and os.path.exists(output_file) and os.path.getsize(output_file) > 0:
                try:
                    from openpyxl import load_workbook
                    from openpyxl.styles import PatternFill, Font, Alignment
                    
                    # Cargar el archivo
                    wb = load_workbook(output_file)
                    
                    # Para cada hoja en el libro
                    for sheet_name in wb.sheetnames:
                        ws = wb[sheet_name]
                        
                        # Aplicar formato a cabeceras
                        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
                        header_font = Font(color="FFFFFF", bold=True)
                        
                        for cell in ws[1]:
                            cell.fill = header_fill
                            cell.font = header_font
                            cell.alignment = Alignment(horizontal="center", vertical="center")
                        
                        # Auto-ajustar ancho de columnas
                        for col in ws.columns:
                            max_length = 0
                            column = col[0].column_letter
                            
                            # Limitar la cantidad de filas a verificar para optimizar
                            row_sample = min(100, ws.max_row)
                            for i in range(1, row_sample + 1):
                                cell = col[i-1]
                                if cell.value:
                                    max_length = max(max_length, len(str(cell.value)))
                            
                            adjusted_width = min(max_length + 2, 40)  # Limitar ancho máximo
                            ws.column_dimensions[column].width = adjusted_width
                    
                    # Guardar el archivo con formato
                    wb.save(output_file)
                    logger.info("Formato aplicado correctamente")
                except Exception as format_error:
                    logger.warning(f"No se pudo aplicar formato avanzado al Excel: {format_error}")
                    # Continuar aunque el formateo falle
            
            # Verificar que el archivo existe y tiene tamaño
            if os.path.exists(output_file) and os.path.getsize(output_file) > 0:
                logger.info(f"Archivo Excel creado correctamente: {os.path.getsize(output_file)} bytes")
                print(f"Archivo Excel creado correctamente: {os.path.getsize(output_file)} bytes")
                
                response = send_file(
                    output_file,
                    mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                    download_name=filename
                )
                
                @response.call_on_close
                def cleanup():
                    try:
                        shutil.rmtree(temp_dir)
                        logger.info("Archivos temporales eliminados")
                    except Exception as e:
                        logger.error(f"Error cleaning up temp files: {e}")
                        
                return response
            else:
                logger.error(f"Error al exportar: excel_ok={excel_ok}, existe={os.path.exists(output_file)}, tamaño={(os.path.getsize(output_file) if os.path.exists(output_file) else 0)} bytes")
                print(f"Error al exportar: excel_ok={excel_ok}, existe={os.path.exists(output_file)}, tamaño={(os.path.getsize(output_file) if os.path.exists(output_file) else 0)} bytes")
                flash('Error al generar el archivo Excel. El archivo parece estar vacío.', 'error')
                return redirect(url_for('search_id'))
                
        except Exception as proc_error:
            logger.error(f"Error al procesar datos para Excel: {str(proc_error)}")
            flash(f'Error al procesar datos para Excel: {str(proc_error)}', 'error')
            return redirect(url_for('search_id'))
            
    except Exception as e:
        logger.error(f"Error en la descarga: {e}", exc_info=True)
        print(f"Error en la descarga: {e}")
        flash(f'Error al generar el archivo: {str(e)}', 'error')
        return redirect(url_for('search_id'))

@app.route('/download_criteria_excel', methods=['POST'])
def download_criteria_excel():
    """Ruta dedicada para la descarga de Excel desde la búsqueda por criterios"""
    logger = logging.getLogger(__name__)
    
    try:
        # Recuperar resultados de la sesión
        results = session.get('search_criteria_results')
        
        # Registra información detallada
        logger.info(f"Exportando resultados. Tipo de datos: {type(results)}")
        print(f"Exportando resultados. Tipo de datos: {type(results)}")
        
        if not results:
            flash('No hay resultados para descargar', 'warning')
            logger.warning("No hay resultados en la sesión para exportar")
            return redirect(url_for('search_criteria'))
            
        # Verificar que hay datos para exportar
        if isinstance(results, list) and len(results) == 0:
            flash('Los resultados están vacíos', 'warning')
            logger.warning("La lista de resultados está vacía")
            return redirect(url_for('search_criteria'))
            
        # Log para depuración
        logger.info(f"Encontrados {len(results)} resultados para exportar")
        print(f"Encontrados {len(results)} resultados para exportar")
        
        # Crear archivo temporal
        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
            temp_path = tmp.name
            
        logger.info(f"Archivo temporal creado: {temp_path}")
        print(f"Archivo temporal creado: {temp_path}")
            
        filename = secure_filename(request.form.get('filename') or 'resultados.xlsx')
        if not filename.endswith('.xlsx'):
            filename += '.xlsx'
            
        # Convertir a DataFrame aquí para garantizar formato correcto
        try:
            # Si los datos son una lista de dicts, convertir a DataFrame
            import pandas as pd
            if isinstance(results, list) and all(isinstance(item, dict) for item in results):
                df = pd.DataFrame(results)
                # Guardar directamente con pandas como respaldo
                logger.info(f"Guardando DataFrame con {len(df)} filas y {len(df.columns)} columnas")
                print(f"Guardando DataFrame con {len(df)} filas y {len(df.columns)} columnas")
                df.to_excel(temp_path, index=False)
                excel_ok = True
            else:
                # Usar el método del manager
                logger.info("Usando manager.export_to_excel para exportar")
                print("Usando manager.export_to_excel para exportar")
                excel_ok = manager.export_to_excel(results, temp_path)
        except Exception as e:
            logger.error(f"Error al convertir datos: {str(e)}")
            print(f"Error al convertir datos: {str(e)}")
            # Intentar el método del manager como último recurso
            excel_ok = manager.export_to_excel(results, temp_path)
            
        if excel_ok and os.path.exists(temp_path) and os.path.getsize(temp_path) > 0:
            logger.info(f"Archivo Excel creado correctamente: {os.path.getsize(temp_path)} bytes")
            print(f"Archivo Excel creado correctamente: {os.path.getsize(temp_path)} bytes")
            
            response = send_file(
                temp_path,
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response.headers['Content-Disposition'] = f'attachment; filename*=UTF-8\'\'{filename}'
            
            @response.call_on_close
            def cleanup():
                try:
                    os.unlink(temp_path)
                    logger.info("Archivo temporal eliminado")
                except Exception as e:
                    logger.error(f"Error al eliminar archivo temporal: {str(e)}")
                    
            return response
        else:
            logger.error(f"Error al exportar: excel_ok={excel_ok}, existe={os.path.exists(temp_path)}, tamaño={(os.path.getsize(temp_path) if os.path.exists(temp_path) else 0)} bytes")
            print(f"Error al exportar: excel_ok={excel_ok}, existe={os.path.exists(temp_path)}, tamaño={(os.path.getsize(temp_path) if os.path.exists(temp_path) else 0)} bytes")
            flash('Error al exportar los resultados. El archivo parece estar vacío.', 'error')
            return redirect(url_for('search_criteria'))
    except Exception as e:
        logger.error(f"Error generating Excel: {e}", exc_info=True)  # Incluir stack trace
        print(f"Error generating Excel: {e}")
        flash(f'Error al generar el archivo: {str(e)}', 'error')
        return redirect(url_for('search_criteria'))

@app.route('/search_criteria', methods=['GET', 'POST'])
def search_criteria():
    search_form = SearchForm()
    results = None
    logger = logging.getLogger(__name__)
    
    if request.method == 'POST' and 'submit_search' in request.form:
        search_type = request.form.get('search_type')
        search_value = request.form.get('search_value')
        
        if search_type == 'list_departments':
            departments = manager.list_departments()
            if departments:
                # Crear lista de diccionarios para poder exportar
                results = [{"Departamento": dept} for dept in departments]
                
                # Guardar resultados en la sesión
                session['search_criteria_results'] = results
        elif search_type == 'list_locations':
            logger.info("Requesting location tree")
            tree_data = manager.location_manager.format_location_tree()
            if tree_data:
                logger.info(f"Found {len(tree_data)} location entries")
                results = [{'text': line} for line in tree_data]
                logger.debug(f"Location data: {results}")
                
                # Guardar resultados en la sesión
                session['search_criteria_results'] = results
            else:
                logger.warning("No locations found")
                flash('No se encontraron ubicaciones', 'warning')
        elif search_type == 'user' and search_value:
            results, message = manager.search_by_user(search_value)
            if not results:
                flash(message, 'warning')
            else:
                # Guardar resultados en la sesión
                session['search_criteria_results'] = results
        elif search_type == 'department' and search_value:
            results, message = manager.search_by_department(search_value)
            
            # Guardar resultados en la sesión
            if results:
                session['search_criteria_results'] = results
            else:
                flash(message, 'warning')
        
        return render_template('search_criteria.html', 
                            search_form=search_form,
                            results=results)
    
    return render_template('search_criteria.html', search_form=search_form)

@app.route('/upload', methods=['GET', 'POST'])
def upload():
    form = FileUploadForm()
    
    if form.validate_on_submit():
        file = form.file.data
        if file:
            # Guardar temporalmente
            temp_file = tempfile.NamedTemporaryFile(delete=False)
            file.save(temp_file.name)
            temp_file.close()
            
            # Leer IDs
            try:
                with open(temp_file.name, 'r') as f:
                    content = f.read().strip()
                
                # Eliminar el archivo temporal
                os.unlink(temp_file.name)
                
                if content:
                    # Guardar en la sesión
                    session['uploaded_ids'] = content
                    flash('IDs cargados correctamente', 'success')
                    return redirect(url_for('search_id'))
                else:
                    flash('El archivo está vacío', 'error')
            except Exception as e:
                flash(f'Error al procesar el archivo: {str(e)}', 'error')
                try:
                    os.unlink(temp_file.name)
                except:
                    pass
                
    return render_template('upload.html', form=form)