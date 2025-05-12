import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
import os
from colorama import Fore, Style, init

class ExcelManager:
    def __init__(self):
        # Inicializar colorama para mostrar colores en consola
        init()
        
    def _get_file_size(self, file_path):
        """Obtiene el tamaño de archivo en formato legible"""
        if not os.path.exists(file_path):
            return "0 B"
            
        size_bytes = os.path.getsize(file_path)
        
        # Formatear tamaño para mejor legibilidad
        if size_bytes < 1024:
            return f"{size_bytes} B"
        elif size_bytes < 1024 * 1024:
            return f"{size_bytes/1024:.2f} KB"
        else:
            return f"{size_bytes/(1024*1024):.2f} MB"

    def export_to_excel(self, df, output_file):
        """Export DataFrame to Excel with formatting"""
        try:
            # Verificar que el DataFrame no esté vacío
            if df is None or df.empty:
                print(f"{Fore.RED}[ERROR] Intentando exportar DataFrame vacío a {output_file}{Style.RESET_ALL}")
                return False
                
            # Verificar que output_file sea válido
            if not output_file:
                print(f"{Fore.RED}[ERROR] Nombre de archivo de salida vacío o inválido{Style.RESET_ALL}")
                return False
                
            # Generar ruta de directorio si no existe
            os.makedirs(os.path.dirname(os.path.abspath(output_file)), exist_ok=True)
                
            # Generate unique filename if needed
            output_file = self._get_unique_filename(output_file)
            print(f"{Fore.CYAN}[INFO] Exportando a archivo: {output_file}{Style.RESET_ALL}")
            
            # Export to Excel - antes de exportar, hacer una copia del DataFrame
            # para evitar modificarlo y evitar problemas de referencias
            import pandas as pd
            export_df = pd.DataFrame(df).copy()
            
            # Print DataFrame info para diagnóstico
            print(f"{Fore.CYAN}[INFO] Información del DataFrame:{Style.RESET_ALL}")
            print(f"{Fore.CYAN}  - Filas: {len(export_df)}{Style.RESET_ALL}")
            print(f"{Fore.CYAN}  - Columnas: {len(export_df.columns)}{Style.RESET_ALL}")
            print(f"{Fore.CYAN}  - Tipos de datos: {export_df.dtypes.to_dict()}{Style.RESET_ALL}")
            
            # Exportar
            export_df.to_excel(output_file, index=False)
            
            # Verificar que el archivo se haya creado y mostrar su tamaño
            if not os.path.exists(output_file):
                print(f"{Fore.RED}[ERROR] El archivo {output_file} no se creó correctamente{Style.RESET_ALL}")
                return False
            
            # Mostrar tamaño del archivo inicial
            file_size = self._get_file_size(output_file)
            print(f"{Fore.CYAN}[INFO] Archivo creado - Tamaño inicial: {file_size}{Style.RESET_ALL}")
                
            # Apply formatting
            if output_file.endswith('.xlsx'):
                try:
                    print(f"{Fore.CYAN}[INFO] Aplicando formato al archivo Excel...{Style.RESET_ALL}")
                    self.format_excel_file(output_file)
                    
                    # Mostrar tamaño del archivo después del formateo
                    new_file_size = self._get_file_size(output_file)
                    print(f"{Fore.CYAN}[INFO] Formato aplicado - Tamaño después de formateo: {new_file_size}{Style.RESET_ALL}")
                    
                    # Mostrar cambio en el tamaño
                    if file_size != new_file_size:
                        print(f"{Fore.CYAN}[INFO] Cambio en tamaño de archivo: {file_size} → {new_file_size}{Style.RESET_ALL}")
                    
                except Exception as e:
                    print(f"{Fore.YELLOW}[ADVERTENCIA] Error al formatear Excel (no crítico): {str(e)}{Style.RESET_ALL}")
                    # Continuar aunque el formateo falle
            
            # Verificación final del archivo
            if os.path.exists(output_file):
                final_size = self._get_file_size(output_file)
                print(f"{Fore.GREEN}[ÉXITO] Exportación a Excel completada exitosamente: {output_file}{Style.RESET_ALL}")
                print(f"{Fore.GREEN}[ÉXITO] Tamaño final del archivo: {final_size}{Style.RESET_ALL}")
                
                # Verificar si el archivo tiene un tamaño no nulo
                if os.path.getsize(output_file) == 0:
                    print(f"{Fore.RED}[ERROR] ¡ADVERTENCIA! El archivo tiene tamaño cero bytes{Style.RESET_ALL}")
                else:
                    print(f"{Fore.GREEN}[ÉXITO] El archivo tiene datos ({final_size}){Style.RESET_ALL}")
                    
                return True
            else:
                print(f"{Fore.RED}[ERROR] El archivo no existe después de la exportación{Style.RESET_ALL}")
                return False
            
        except Exception as e:
            print(f"{Fore.RED}[ERROR] Error al exportar a Excel: {str(e)}{Style.RESET_ALL}")
            import traceback
            print(f"{Fore.RED}[ERROR] Detalle del error:\n{traceback.format_exc()}{Style.RESET_ALL}")
            return False

    def _get_unique_filename(self, output_file):
        """Ensure the filename is unique by adding a numeric suffix if needed"""
        if not os.path.exists(output_file):
            return output_file
            
        # Si el archivo ya existe, agregar un sufijo numérico
        base, ext = os.path.splitext(output_file)
        counter = 1
        while os.path.exists(f"{base}_{counter}{ext}"):
            counter += 1
        return f"{base}_{counter}{ext}"

    def format_excel_file(self, file_path):
        """Apply formatting to Excel file"""
        # Tamaño antes de formatear
        initial_size = self._get_file_size(file_path)
        print(f"{Fore.CYAN}[INFO] Iniciando formateo - Tamaño actual: {initial_size}{Style.RESET_ALL}")
        
        # Cargar el archivo
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active

        # Personalizar estilos
        header_fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        row_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        
        # Aplicar formato automático a columnas
        print(f"{Fore.CYAN}[INFO] Ajustando ancho de columnas...{Style.RESET_ALL}")
        self._auto_adjust_columns(ws)
        
        # Aplicar estilos mejorados
        print(f"{Fore.CYAN}[INFO] Aplicando estilos a encabezados...{Style.RESET_ALL}")
        self._apply_header_styles(ws, header_fill, header_font)
        
        print(f"{Fore.CYAN}[INFO] Aplicando estilos a filas...{Style.RESET_ALL}")
        self._apply_row_styles(ws, row_fill)

        # Guardar el archivo
        wb.save(file_path)
        
        # Tamaño después de formatear
        final_size = self._get_file_size(file_path)
        print(f"{Fore.CYAN}[INFO] Formateo completado - Tamaño después: {final_size}{Style.RESET_ALL}")
        
        return True

    def _auto_adjust_columns(self, ws):
        """Auto-adjust column widths"""
        column_count = 0
        for col in ws.columns:
            column_count += 1
            max_length = 0
            col_letter = col[0].column_letter
            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = min(max_length + 3, 50)  # Cap width at 50
            ws.column_dimensions[col_letter].width = adjusted_width
        
        print(f"{Fore.CYAN}[INFO] Ajustadas {column_count} columnas en el Excel{Style.RESET_ALL}")
            
    def _apply_header_styles(self, ws, header_fill, header_font):
        """Apply styles to header row"""
        header_count = 0
        for cell in ws[1]:
            header_count += 1
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        print(f"{Fore.CYAN}[INFO] Aplicado estilo a {header_count} celdas de encabezado{Style.RESET_ALL}")
            
    def _apply_row_styles(self, ws, row_fill):
        """Apply alternating row styles"""
        styled_rows = 0
        for row_idx, row in enumerate(ws.iter_rows(min_row=2)):
            # Aplicar estilo alternado a filas
            if row_idx % 2 == 0:
                styled_rows += 1
                for cell in row:
                    cell.fill = row_fill
                    
        print(f"{Fore.CYAN}[INFO] Aplicado estilo a {styled_rows} filas de datos{Style.RESET_ALL}")